"""Deterministic extraction: ACES AMQ xlsx -> rule cards JSON. No LLM anywhere in this step.

Outputs:
  data/cards_all.json     every question card, all families/questionnaires, scope-classified
  data/cards_base.json    the locked base: post-closing, Fannie-cut, live
  data/extract_summary.json
"""
import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / 'source' / 'amqs-sept-2025-retail.xlsx'
HDR = ['qname', 'cat', 'anscat', 'excname', 'qcode', 'qtext', 'qresp',
       'qcrit', 'exccode', 'sig', 'excdesc', 'aor1', 'aor2', 'qcritq']

GOV_PREFIXES = ('O-FHA', 'O-VA', 'O-RHS')
FREDDIE_PREFIXES = ('O-FRD', 'FRD-Closing')
PROGRAM_FAMILIES = ('SONYMA', 'O-SONYMA', 'Portfolio', 'Overlay', 'Medical Professional')


def clean(v):
    if v is None:
        return None
    s = str(v).replace('_x000D_', '\n').replace('\xa0', ' ').strip()
    return s if s and s != 'None' else None


def family(code):
    return re.sub(r'[-\s]*\d+$', '', str(code or '')).rstrip('-').strip()


def scope_status(fam, cat):
    if cat == 'Discarded':
        return 'discarded'
    if 'COVID' in fam:
        return 'excluded_covid'
    if fam.startswith(GOV_PREFIXES):
        return 'excluded_government'
    if fam.startswith(FREDDIE_PREFIXES):
        return 'excluded_freddie'
    if any(fam.startswith(p) for p in PROGRAM_FAMILIES):
        return 'excluded_program'
    return 'base'


def main():
    wb = load_workbook(SRC, read_only=True, data_only=True)
    ws = wb['Report 1']
    rows = [dict(zip(HDR, r)) for r in ws.iter_rows(min_row=5, values_only=True) if any(r)]

    grouped = defaultdict(list)
    for r in rows:
        qn = 'post_closing' if 'Post-Closing' in str(r['qname']) else 'pre_funding'
        grouped[(qn, str(r['qcode']))].append(r)

    cards = []
    for (qn, qcode), rs in grouped.items():
        # A question can carry retired answer-options under the 'Discarded' category
        # alongside live rows; the card is live iff any row is non-Discarded, and only
        # live rows contribute answer options.
        live_rs = [r for r in rs if str(r['cat']) != 'Discarded']
        n_retired = len(rs) - len(live_rs)
        if live_rs:
            rs = live_rs
        r0 = rs[0]
        fam = family(qcode)
        cat = clean(r0['cat']) or ''
        card = {
            'card_id': f"{'PC' if qn == 'post_closing' else 'PF'}::{qcode}",
            'retired_options_dropped': n_retired,
            'questionnaire': qn,
            'question_code': qcode,
            'family': fam,
            'category': cat,
            'question_text': clean(r0['qtext']),
            'applicability_sql': clean(r0['qcrit']),
            'skip_logic': clean(r0['qcritq']),
            'scope': {'status': scope_status(fam, cat), 'route': 'FNM'},
            'answer_options': [
                {
                    'response': clean(r['qresp']),
                    'answer_exception_name': clean(r['excname']),
                    'answer_category_criteria': clean(r['anscat']),
                    'exception_code': clean(r['exccode']),
                    'significance': clean(r['sig']),
                    'exception_description': clean(r['excdesc']),
                    'aor_1': clean(r['aor1']),
                    'aor_2': clean(r['aor2']),
                }
                for r in rs
            ],
        }
        cards.append(card)

    cards.sort(key=lambda c: (c['questionnaire'], c['family'], c['question_code']))
    base = [c for c in cards if c['questionnaire'] == 'post_closing' and c['scope']['status'] == 'base']

    summary = {
        'total_rows': len(rows),
        'total_cards': len(cards),
        'base_cards_pc_fannie_cut': len(base),
        'base_defect_options': sum(1 for c in base for o in c['answer_options'] if o['exception_code']),
        'by_status': {},
        'base_by_category': {},
        'base_by_family': {},
    }
    st = defaultdict(int)
    for c in cards:
        st[f"{c['questionnaire']}::{c['scope']['status']}"] += 1
    summary['by_status'] = dict(sorted(st.items()))
    bc, bf = defaultdict(int), defaultdict(int)
    for c in base:
        bc[c['category']] += 1
        bf[c['family']] += 1
    summary['base_by_category'] = dict(sorted(bc.items(), key=lambda kv: -kv[1]))
    summary['base_by_family'] = dict(sorted(bf.items(), key=lambda kv: -kv[1]))

    (ROOT / 'data').mkdir(exist_ok=True)
    (ROOT / 'data' / 'cards_all.json').write_text(json.dumps(cards, indent=1))
    (ROOT / 'data' / 'cards_base.json').write_text(json.dumps(base, indent=1))
    (ROOT / 'data' / 'extract_summary.json').write_text(json.dumps(summary, indent=2))
    print(json.dumps(summary, indent=2))


if __name__ == '__main__':
    main()
