"""
Compile-metadata wiring -- three small, additive fixes bundled together:

  FIX A: Check.question_code persists the AMQ "Question Code" onto the
         compiled Check (compile_llm.compile_row()), audit metadata only.
  FIX B: taxonomy.load_rows() captures each row's real workbook locator
         ("sheet" + "source_row"), and sample.stratified_sample() derives a
         stable row_id from that locator instead of the random-sample
         enumeration index.
  FIX C: assemble_ruleset() wires CompiledCheckDraft.grounding and
         .source_locator onto the signed Ruleset's RuleIntentRecord --
         previously computed per-check but silently dropped before the
         signed artifact existed.

Run from p0/:  python -m pytest tests/test_compile_metadata_wiring.py -v
Python 3.9 compatible.
"""
from __future__ import annotations

import os
import sys
import tempfile

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from qc_engine.ruleset import Check
from qc_engine.compiler import compile_llm as C
from qc_engine.compiler import sample as S
from eval_synth import taxonomy as T


# --- FIX A: Check.question_code --------------------------------------------

def test_check_question_code_round_trips_through_to_dict():
    check = Check(
        id="chk-001", name="Final URLA signed", field_name="urla_signed",
        kind="predicate", severity="CRITICAL", predicate="is_true",
        question_code="Final URLA",
    )
    assert check.question_code == "Final URLA"
    as_dict = check.to_dict()
    assert as_dict["question_code"] == "Final URLA", (
        f"expected question_code to round-trip through to_dict(), got {as_dict}")


def test_check_question_code_defaults_to_none_when_unset():
    # Existing call sites that don't pass question_code must be unaffected.
    check = Check(id="chk-002", name="x", field_name="x", kind="predicate",
                  severity="INFO")
    assert check.question_code is None
    assert check.to_dict()["question_code"] is None


# --- FIX B (taxonomy.load_rows): sheet + source_row capture ----------------

def _build_test_workbook(path: str) -> None:
    """A minimal, real .xlsx built inline (not a checked-in binary fixture)
    matching the AMQ export's _STANDARD_COLS layout: rows 1-4 are header/
    metadata (skipped by load_rows' min_row=5), data starts at real Excel
    row 5. Two sheets, so "sheet" capture is actually exercised across more
    than one worksheet -- 010a's own "every sheet, not just the first" fix
    this module already relies on."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"

    def _write_row(ws, row_num, qcode, defect_text, exc_code, significance):
        # index (0-based, matches taxonomy._STANDARD_COLS) -> openpyxl column
        # (1-based) is index + 1.
        ws.cell(row=row_num, column=1, value="Test AMQ Questionnaire")  # r[0]
        ws.cell(row=row_num, column=2, value="Test Category")           # r[1] category
        ws.cell(row=row_num, column=5, value=qcode)                     # r[4] qcode
        ws.cell(row=row_num, column=7, value=defect_text)               # r[6] defect_text
        ws.cell(row=row_num, column=8, value="")                        # r[7] sql_criteria
        ws.cell(row=row_num, column=9, value=exc_code)                  # r[8] exception_code
        ws.cell(row=row_num, column=10, value=significance)             # r[9] significance

    _write_row(ws1, 5, "Q100", "Signature is missing on document", "TEST-001", "Critical")
    _write_row(ws1, 6, "Q101", "LTV exceeds 80 percent", "TEST-002", "Major")

    ws2 = wb.create_sheet("Sheet2")
    _write_row(ws2, 5, "Q200", "Document is not present in the file", "TEST-003", "Critical")

    wb.save(path)


def test_load_rows_captures_sheet_and_source_row():
    with tempfile.TemporaryDirectory() as tmpdir:
        path = os.path.join(tmpdir, "test_amq.xlsx")
        _build_test_workbook(path)
        rows = T.load_rows(path)

    assert len(rows) == 3, f"expected 3 real defect rows, got {len(rows)}: {rows}"
    by_exc = {r["exception_code"]: r for r in rows}

    row1 = by_exc["TEST-001"]
    assert row1["sheet"] == "Sheet1", f"expected sheet='Sheet1', got {row1!r}"
    assert row1["source_row"] == 5, f"expected source_row=5, got {row1!r}"

    row2 = by_exc["TEST-002"]
    assert row2["sheet"] == "Sheet1", f"expected sheet='Sheet1', got {row2!r}"
    assert row2["source_row"] == 6, f"expected source_row=6, got {row2!r}"

    row3 = by_exc["TEST-003"]
    assert row3["sheet"] == "Sheet2", f"expected sheet='Sheet2', got {row3!r}"
    assert row3["source_row"] == 5, f"expected source_row=5, got {row3!r}"

    # Purely additive: every pre-existing key is still present and correctly
    # shaped -- this fix must not have touched them.
    for r in rows:
        for key in ("category", "qcode", "defect_text", "sql_criteria",
                    "exception_code", "significance"):
            assert key in r, f"pre-existing key {key!r} missing from row: {r}"


# --- FIX B (sample.stratified_sample): stable row_id from the real locator -

def _synthetic_pool_row(source_file, sheet, source_row, defect_text="x is missing"):
    return {
        "category": "Test Category", "qcode": "Q1", "defect_text": defect_text,
        "sql_criteria": "", "exception_code": "TEST-999", "significance": "Critical",
        "sheet": sheet, "source_row": source_row,
        "archetype_id": "MISSING", "engine_kind": "predicate",
        "source_file": source_file,
    }


def test_stratified_sample_row_id_stable_across_seeds_when_locator_present(monkeypatch):
    # A small, fixed pool of rows carrying a real (source_file, sheet,
    # source_row) locator each -- the same shape load_rows()/
    # classify_all_rows() now produce.
    pool = [_synthetic_pool_row("book.xlsx", "Sheet1", 5 + i) for i in range(8)]
    monkeypatch.setattr(S, "classify_all_rows", lambda: {"predicate": pool})

    sample_a = S.stratified_sample(n_total=8, seed=111,
                                    mix_ratios={"predicate": 1.0})
    sample_b = S.stratified_sample(n_total=5, seed=222,
                                    mix_ratios={"predicate": 1.0})

    ids_a_by_locator = {(r["sheet"], r["source_row"]): r["row_id"] for r in sample_a}
    ids_b_by_locator = {(r["sheet"], r["source_row"]): r["row_id"] for r in sample_b}

    assert ids_a_by_locator, "sample_a produced no rows"
    for locator, row_id in ids_a_by_locator.items():
        expected = f"book.xlsx:{locator[0]}:{locator[1]}"
        assert row_id == expected, (
            f"row_id for locator {locator} was {row_id!r}, expected {expected!r}")

    # Any locator drawn into BOTH differently-seeded/differently-sized
    # samples must get the SAME row_id -- proving it's derived from the
    # row's own identity, not from its position in that particular draw
    # (the old f"{bucket}-{i:03d}" scheme was NOT stable this way).
    shared = set(ids_a_by_locator) & set(ids_b_by_locator)
    assert shared, "test setup should guarantee overlap for a meaningful assertion"
    for locator in shared:
        assert ids_a_by_locator[locator] == ids_b_by_locator[locator], (
            f"row_id for {locator} differed across seeds: "
            f"{ids_a_by_locator[locator]!r} vs {ids_b_by_locator[locator]!r}")


def test_stratified_sample_falls_back_to_bucket_index_id_without_locator(monkeypatch):
    # A row with no sheet/source_row (e.g. a hand-authored synthetic row
    # predating this fix) must still get an id -- the old bucket-index
    # scheme, unchanged.
    pool = [{
        "category": "Test Category", "qcode": "Q1", "defect_text": "x is missing",
        "sql_criteria": "", "exception_code": "TEST-999", "significance": "Critical",
        "archetype_id": "MISSING", "engine_kind": "predicate",
        "source_file": "book.xlsx",
    } for _ in range(3)]
    monkeypatch.setattr(S, "classify_all_rows", lambda: {"predicate": pool})

    sample = S.stratified_sample(n_total=3, seed=1, mix_ratios={"predicate": 1.0})
    assert len(sample) == 3
    for i, row in enumerate(sample):
        assert row["row_id"] == f"predicate-{i:03d}", (
            f"expected fallback bucket-index row_id, got {row['row_id']!r}")


# --- FIX C: assemble_ruleset() wires grounding + source_locator onto the ---
# --- signed Ruleset's RuleIntentRecord --------------------------------------

def _grounded_draft(row_id, check_id, source_locator):
    check = Check(id=check_id, name="Gift funds documented",
                  field_name="gift_funds_source_documented", kind="predicate",
                  severity="CRITICAL", predicate="is_true")
    return C.CompiledCheckDraft(
        row_id=row_id, check=check, source_text="Gift funds not documented",
        extracted_intent="Fails when gift funds source isn't documented.",
        grounding=C.GroundingRecord(
            kb_program="FHA", kb_version=3, section_ids=["sec-1", "sec-2"]),
        source_locator=source_locator,
    )


def test_assemble_ruleset_carries_source_locator_and_grounding_onto_intent_records():
    drafts = [
        _grounded_draft("row-a", "chk-a", "book.xlsx:Sheet1:5"),
        _grounded_draft("row-b", "chk-b", "book.xlsx:Sheet2:9"),
    ]
    rs = C.assemble_ruleset(drafts, ruleset_id="batch-metadata-wiring", version=1,
                            signed_by="test", signed_at="2026-07-29T00:00:00Z")

    assert len(rs.intent_records) == 2
    rec_a = rs.intent_for("chk-a")
    assert rec_a is not None
    assert rec_a.source_locator == "book.xlsx:Sheet1:5", (
        f"expected source_locator to be wired through, got {rec_a.source_locator!r}")
    assert rec_a.kb_program == "FHA", f"expected kb_program='FHA', got {rec_a.kb_program!r}"
    assert rec_a.kb_version == 3, f"expected kb_version=3, got {rec_a.kb_version!r}"
    assert rec_a.section_ids == ["sec-1", "sec-2"], (
        f"expected section_ids to round-trip, got {rec_a.section_ids!r}")

    rec_b = rs.intent_for("chk-b")
    assert rec_b.source_locator == "book.xlsx:Sheet2:9"


def test_assemble_ruleset_intent_record_grounding_fields_none_when_ungrounded():
    # 002c FR-006 precedent: grounding is additive, never a hard blocker --
    # a draft with grounding=None must produce an intent_record with the new
    # fields cleanly None, not a crash or a stray sentinel.
    check = Check(id="chk-ungrounded", name="x", field_name="x", kind="predicate",
                  severity="INFO", predicate="is_true")
    draft = C.CompiledCheckDraft(
        row_id="row-c", check=check, source_text="x", extracted_intent="x",
        grounding=None, source_locator=None,
    )
    rs = C.assemble_ruleset([draft], ruleset_id="batch-ungrounded", version=1,
                            signed_by="test", signed_at="2026-07-29T00:00:00Z")
    rec = rs.intent_for("chk-ungrounded")
    assert rec.source_locator is None
    assert rec.kb_program is None
    assert rec.kb_version is None
    assert rec.section_ids is None


if __name__ == "__main__":
    import pytest
    raise SystemExit(pytest.main([__file__, "-v"]))
