"""
Exports the 495 UNSPECIFIED-threshold check instances (from the comprehensive
8,399-check ruleset) to a spreadsheet for SME eyeball review -- one row per
instance (not deduplicated), sorted by check_id so duplicate instances of the
same real-world rule sit next to each other for easy comparison.

Python 3.9 compatible.
"""
from __future__ import annotations

import json
import os
from collections import Counter

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
RESULT_DIR = os.path.normpath(os.path.join(HERE, "..", "..", "..", "result"))
RULESET_PATH = os.path.join(RESULT_DIR, "rules", "comprehensive_ruleset.json")
APPLICABILITY_PATH = os.path.join(RESULT_DIR, "rules", "comprehensive_applicability.json")
OUT_PATH = os.path.join(RESULT_DIR, "..", "output", "UNSPECIFIED-thresholds-for-SME-review.xlsx")


def main() -> None:
    data = json.load(open(RULESET_PATH))
    applicability = json.load(open(APPLICABILITY_PATH))
    checks = data["content"]["checks"]

    unspecified = [c for c in checks
                   if c.get("threshold") == "UNSPECIFIED" or c.get("tolerance") == "UNSPECIFIED"]
    unspecified.sort(key=lambda c: c["id"])

    group_sizes = Counter(c["id"] for c in unspecified)
    seen_so_far: Counter = Counter()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UNSPECIFIED thresholds"

    headers = [
        "check_id", "instance # of group", "check_name", "kind",
        "field_name", "missing_value", "operator", "severity",
        "program(s)", "message_fail (context for the real number)",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4A4E8F")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    for c in unspecified:
        seen_so_far[c["id"]] += 1
        missing_value = "threshold" if c.get("threshold") == "UNSPECIFIED" else "tolerance"
        programs = ", ".join(applicability.get(c["id"], ["UNTAGGED"]))
        ws.append([
            c["id"],
            f"{seen_so_far[c['id']]} of {group_sizes[c['id']]}",
            c.get("name", ""),
            c.get("kind", ""),
            c.get("field_name", ""),
            missing_value,
            c.get("operator", ""),
            c.get("severity", ""),
            programs,
            c.get("message_fail", ""),
        ])

    widths = [34, 14, 42, 16, 30, 12, 9, 10, 26, 70]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column_letter == "J"))

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Total UNSPECIFIED instances", len(unspecified)])
    ws2.append(["Unique check ids involved", len(group_sizes)])
    ws2.append(["Source ruleset", "result/rules/comprehensive_ruleset.json"])
    ws2.append(["Generated", "2026-07-23"])
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 50
    for r in range(1, 5):
        ws2.cell(row=r, column=1).font = Font(bold=True)

    wb.save(OUT_PATH)
    print(f"[written] {OUT_PATH}")
    print(f"  {len(unspecified)} rows, {len(group_sizes)} unique check ids")


if __name__ == "__main__":
    main()
