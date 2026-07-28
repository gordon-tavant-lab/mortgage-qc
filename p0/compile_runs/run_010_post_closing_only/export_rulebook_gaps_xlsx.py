"""
Exports the 37 "rulebook gap" checks (UNSPECIFIED_THRESHOLD -- loan 01,
result/rules/post_closing_only_ruleset.json) to a spreadsheet, with every
original column copied verbatim from the client's real source workbook row
each check was compiled from -- not just the exception code, the FULL row
(Questionnaire Name, Question Category Name, Question Text, Question
Response, Question Criteria, Exception Code, Default Significance,
Exception Description, etc.), read directly from demo/rules/*.xlsx using
the exact physical row `post_closing_only_provenance.json` points to.

One output row per (check, source-row) pair -- a few checks were compiled
from more than one source row (the same real-world condition restated
across program sheets), so those get more than one row here, each showing
which specific source row it came from.

Python 3.9 compatible.
"""
from __future__ import annotations

import json
import os

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
P0 = os.path.dirname(os.path.dirname(HERE))
RULES_DIR = os.path.normpath(os.path.join(P0, "..", "demo", "rules"))
RESULT_DIR = os.path.normpath(os.path.join(P0, "..", "result"))
LOAN01_RESULT_PATH = os.path.join(HERE, "loan01_with_provenance.json")
RULESET_PATH = os.path.join(RESULT_DIR, "rules", "post_closing_only_ruleset.json")
OUT_PATH = os.path.join(P0, "..", "output", "LOAN-01-RULEBOOK-GAPS-SOURCE-ROWS-2026-07-24.xlsx")

# Union of every column header seen across both source workbooks' sheets, in
# a fixed display order -- a sheet missing a given column just leaves it blank.
UNION_HEADERS = [
    "Questionnaire Name", "Question Category Name", "Question Answers Category Criteria",
    "Question Answers Exception Name", "Question Code", "Question Text",
    "Question Response", "Question Criteria", "Exception Code",
    "Default Significance", "Exception Description", "Default AOR 1", "Default AOR 2",
]

# Each sheet's own header row number (confirmed by direct inspection --
# "Post Closing Oct 2025" has no leading title/date rows; "Report 1" does).
SHEET_HEADER_ROW = {
    ("Private Bank Oct 2025 PC and Nov 2025 PF.xlsx", "Post Closing Oct 2025"): 1,
    ("PF and PC Sept 2025 AMQs - Retail.xlsx", "Report 1"): 4,
}
# provenance's source_row is 0-indexed from openpyxl's iter_rows(min_row=5, ...)
# -- physical sheet row = 5 + source_row, regardless of that sheet's own
# header position (the loader used min_row=5 uniformly).
PROVENANCE_MIN_ROW = 5


def _sheet_row_to_dict(ws, header_row: int, physical_row: int) -> dict:
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, len(UNION_HEADERS) + 2)]
    values = [ws.cell(row=physical_row, column=c).value for c in range(1, len(UNION_HEADERS) + 2)]
    row_dict = {}
    for h, v in zip(headers, values):
        if h:
            row_dict[str(h).strip()] = v
    return row_dict


def main() -> None:
    result = json.load(open(LOAN01_RESULT_PATH))
    gaps = [r for r in result["surfaced_results"] if r["review_reason"] == "UNSPECIFIED_THRESHOLD"]

    ruleset = json.load(open(RULESET_PATH))
    id_to_field = {}
    for c in ruleset["content"]["checks"]:
        id_to_field.setdefault(c["id"], c.get("field_name"))

    # Open each referenced workbook once, cache worksheets.
    open_wbs = {}
    def get_ws(source_file, sheet_name):
        key = (source_file, sheet_name)
        if source_file not in open_wbs:
            open_wbs[source_file] = openpyxl.load_workbook(
                os.path.join(RULES_DIR, source_file), read_only=True, data_only=True)
        return open_wbs[source_file][sheet_name]

    wb = openpyxl.Workbook()
    ws_out = wb.active
    ws_out.title = "Rulebook Gaps (Loan 01)"

    headers = ["check_id", "missing_field", "tool_message", "source_file", "sheet_name",
               "excel_row_#"] + UNION_HEADERS
    ws_out.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws_out.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4A4E8F")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws_out.freeze_panes = "A2"

    row_count = 0
    for r in gaps:
        check_id = r["check_id"]
        field = id_to_field.get(check_id, "")
        for p in r["provenance"]:
            key = (p["source_file"], p["sheet_name"])
            header_row = SHEET_HEADER_ROW.get(key)
            ws = get_ws(p["source_file"], p["sheet_name"])
            physical_row = PROVENANCE_MIN_ROW + p["source_row"]
            row_data = _sheet_row_to_dict(ws, header_row, physical_row) if header_row else {}
            out_row = [check_id, field, r["message"], p["source_file"], p["sheet_name"], physical_row]
            out_row += [row_data.get(h, "") for h in UNION_HEADERS]
            ws_out.append(out_row)
            row_count += 1

    widths = [34, 20, 46, 32, 20, 10] + [26, 20, 22, 22, 20, 40, 40, 26, 16, 40, 14, 14]
    for i, w in enumerate(widths, start=1):
        if i <= ws_out.max_column:
            ws_out.column_dimensions[get_column_letter(i)].width = w

    for row in ws_out.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Rulebook gap checks (loan 01)", len(gaps)])
    ws2.append(["Total rows (checks x source rows)", row_count])
    ws2.append(["Source ruleset", "result/rules/post_closing_only_ruleset.json"])
    ws2.append(["Source workbooks", "demo/rules/*.xlsx (client's real AMQ rulebook)"])
    ws2.append(["Generated", "2026-07-24"])
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 60
    for r in range(1, 6):
        ws2.cell(row=r, column=1).font = Font(bold=True)

    wb.save(OUT_PATH)
    for wbx in open_wbs.values():
        wbx.close()
    print(f"[written] {OUT_PATH}")
    print(f"  {len(gaps)} unique gap checks, {row_count} total rows (some checks cite >1 source row)")


if __name__ == "__main__":
    main()
